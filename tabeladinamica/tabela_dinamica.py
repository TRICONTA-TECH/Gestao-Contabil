import pandas as pd
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font

# configuração dos arquivos
INPUT = "gestta.busca.xlsx"
OUTPUT = "saida_filtrada.xlsx"

# carregar planilha
df = pd.read_excel(INPUT)

# seleciona colunas relevantes
df = df[["Responsável", "Cliente", "Competência", "Status"]].copy()

# normalizar status
def norm_status(s):
    if pd.isna(s):
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode()
    return s.upper().strip()

df["Status"] = df["Status"].apply(norm_status)

# pivot
pivot = pd.pivot_table(
    df,
    index=["Responsável", "Cliente", "Competência"],
    columns="Status",
    aggfunc="size",
    fill_value=0
).reset_index()

# garantir colunas
for c in ("ABERTO", "CONCLUIDO"):
    if c not in pivot.columns:
        pivot[c] = 0

# total geral por cliente
pivot["Total Geral"] = pivot.groupby(["Responsável", "Cliente"])["Competência"].transform("count")

agg_df = pivot[["Responsável", "Cliente", "Competência", "ABERTO", "CONCLUIDO", "Total Geral"]].copy()

final_rows = []

subtotal_indices = []  # linhas para grifar

for responsavel, grp_resp in agg_df.groupby("Responsável", sort=False):
    grp_resp = grp_resp.sort_values(["Cliente", "Competência"], kind="stable")
    primeira_linha_resp = True

    for cliente, grp_cli in grp_resp.groupby("Cliente", sort=False):
        primeira_linha_cli = True

        for _, row in grp_cli.iterrows():
            final_rows.append({
                "Responsável": responsavel if primeira_linha_resp else "",
                "Cliente": cliente if primeira_linha_cli else "",
                "Competência": row["Competência"],
                "Status ABERTO": int(row.get("ABERTO", 0)) if row.get("ABERTO", 0) != 0 else "",
                "Status CONCLUÍDO": int(row.get("CONCLUIDO", 0)),
                "Total Geral": int(row["Total Geral"])
            })
            primeira_linha_resp = False
            primeira_linha_cli = False

        # subtotal cliente
        total_comp_cli = len(grp_cli)
        soma_aberto_cli = int(grp_cli["ABERTO"].sum())
        soma_concluido_cli = int(grp_cli["CONCLUIDO"].sum())
        final_rows.append({
            "Responsável": "",
            "Cliente": cliente,
            "Competência": total_comp_cli,
            "Status ABERTO": soma_aberto_cli if soma_aberto_cli != 0 else "",
            "Status CONCLUÍDO": soma_concluido_cli,
            "Total Geral": total_comp_cli
        })
        subtotal_indices.append(len(final_rows))  # marcar linha

    # subtotal responsável
    total_comp_resp = len(grp_resp)
    soma_aberto_resp = int(grp_resp["ABERTO"].sum())
    soma_concluido_resp = int(grp_resp["CONCLUIDO"].sum())
    final_rows.append({
        "Responsável": "",
        "Cliente": "",
        "Competência": total_comp_resp,
        "Status ABERTO": soma_aberto_resp if soma_aberto_resp != 0 else "",
        "Status CONCLUÍDO": soma_concluido_resp,
        "Total Geral": total_comp_resp
    })
    subtotal_indices.append(len(final_rows))  # marcar linha

# total sistema
total_sistema = len(agg_df)
total_aberto = int(agg_df["ABERTO"].sum())
total_concluido = int(agg_df["CONCLUIDO"].sum())
final_rows.append({
    "Responsável": "",
    "Cliente": "",
    "Competência": total_sistema,
    "Status ABERTO": total_aberto if total_aberto != 0 else "",
    "Status CONCLUÍDO": total_concluido,
    "Total Geral": total_sistema
})
subtotal_indices.append(len(final_rows))  # marcar linha

# montar DataFrame
final_df = pd.DataFrame(final_rows, columns=[
    "Responsável", "Cliente", "Competência",
    "Status ABERTO", "Status CONCLUÍDO", "Total Geral"
])

final_df[["Status CONCLUÍDO", "Total Geral"]] = \
    final_df[["Status CONCLUÍDO", "Total Geral"]].astype(int)

# exportar para excel
final_df.to_excel(OUTPUT, index=False)

# FORMATAÇÃO com openpyxl
wb = load_workbook(OUTPUT)
ws = wb.active
bold_font = Font(bold=True)

for idx in subtotal_indices:
    row_idx = idx + 1  # +1 porque openpyxl é 1-indexed e tem header
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col).font = bold_font

wb.save(OUTPUT)

